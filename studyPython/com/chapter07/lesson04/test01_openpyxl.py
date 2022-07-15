from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#定义Workbook实例化
wb = Workbook()
#定义文件名
dest_filename = 'empty_book.xlsx'

#定义正在运行的工作表ws1并把range names赋值给ws1的title
ws1 = wb.active
ws1.title = "range names"

#创建39*599的数据
for row in range(1, 40):
    ws1.append(range(600))

#创建sheet页签并给某个单元格写值
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14

#创建sheet页签并批量给单元格挟制
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

#保存表数据
wb.save(filename = dest_filename)